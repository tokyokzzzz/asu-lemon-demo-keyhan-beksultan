import json
import os
from datetime import datetime
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RECORDS_FILE = "uploads/records.json"

# Expert evaluation columns with their max scores
EXPERT_COLUMNS = [
    ("Стратегическая релевантность (20%)", "strategic_relevance", 20),
    ("Цель и задачи (10%)", "goals_and_tasks", 10),
    ("Научная новизна (15%)", "scientific_novelty", 15),
    ("Практическая применимость (20%)", "practical_applicability", 20),
    ("Ожидаемые результаты (15%)", "expected_results", 15),
    ("Соц-экономический эффект (10%)", "socioeconomic_impact", 10),
    ("Реализуемость (10%)", "feasibility", 10),
]


def _load_records() -> List[Dict[str, Any]]:
    if not os.path.exists(RECORDS_FILE):
        return []
    try:
        with open(RECORDS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def _save_records(records: List[Dict[str, Any]]) -> None:
    os.makedirs("uploads", exist_ok=True)
    with open(RECORDS_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


def add_score_record(
    filename: str,
    original_score: int,
    language: str,
    corrected_score: Optional[int] = None,
    score_breakdown: Optional[Dict[str, int]] = None,
) -> Dict[str, Any]:
    records = _load_records()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    record = {
        "filename": filename,
        "original_score": original_score,
        "corrected_score": corrected_score,
        "language": language,
        "timestamp": timestamp,
        "score_breakdown": score_breakdown or {},
        "organization": "",
        "expert": "",
        "comment": "",
    }

    records.append(record)
    _save_records(records)
    return record


def update_corrected_score(filename: str, corrected_score: int) -> bool:
    records = _load_records()
    for i in range(len(records) - 1, -1, -1):
        if records[i]["filename"] == filename:
            records[i]["corrected_score"] = corrected_score
            _save_records(records)
            return True
    return False


def export_to_excel() -> str:
    records = _load_records()

    wb = Workbook()
    ws = wb.active
    ws.title = "Экспертная оценка"

    # Header row
    headers = ["№", "Название ТЗ", "Организация", "Эксперт"]
    headers += [col[0] for col in EXPERT_COLUMNS]
    headers += ["Итоговый балл", "Комментарий эксперта"]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 45

    # Data rows
    for row_num, record in enumerate(records, 2):
        breakdown = record.get("score_breakdown", {})
        total_score = record.get("original_score", 0)

        row_data = [
            row_num - 1,
            record.get("filename", ""),
            record.get("organization", ""),
            record.get("expert", ""),
        ]
        row_data += [breakdown.get(col[1], "") for col in EXPERT_COLUMNS]
        row_data += [total_score, record.get("comment", "")]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Color total score cell
        score_col = len(headers)  # last column before comment
        score_col_idx = headers.index("Итоговый балл") + 1
        score_cell = ws.cell(row=row_num, column=score_col_idx)
        if total_score < 50:
            score_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            score_cell.font = Font(color="FFFFFF", bold=True)
        elif total_score < 75:
            score_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            score_cell.font = Font(color="000000", bold=True)
        else:
            score_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            score_cell.font = Font(color="FFFFFF", bold=True)

        # Alternate row shading
        if row_num % 2 == 0:
            row_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
            for col_num in range(1, len(headers) + 1):
                c = ws.cell(row=row_num, column=col_num)
                if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    c.fill = row_fill

    # Column widths
    col_widths = [5, 35, 20, 20] + [14] * len(EXPERT_COLUMNS) + [14, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"uploads/expert_scores_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path


def get_all_records() -> List[Dict[str, Any]]:
    return _load_records()


def clear_records() -> None:
    _save_records([])
